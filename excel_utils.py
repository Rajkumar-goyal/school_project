import openpyxl
import io
import sqlite3
from werkzeug.security import generate_password_hash
from database import get_db_connection
from models import Student, User, Subject, Class, Result

#
# >>> FIX:
# I removed the incorrect "@app.route(...)" line that was here.
# This file should only contain the class definition.
#

class ExcelImporter:

    @staticmethod
    def download_template(template_type):
        """
        Generates an Excel template file in memory for a given data type.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = []
        filename = "template.xlsx"

        if template_type == 'users':
            headers = ['username', 'password', 'role', 'name', 'email']
            ws.title = "Users"
            filename = "users_template.xlsx"
        
        elif template_type == 'students':
            headers = [
                'full_name', 'email', 'gender', 'date_of_birth', 
                'class_name', 'section', 'roll_number', 
                'fathers_name', 'mothers_name', 'mobile_number', 'academic_year'
            ]
            ws.title = "Students"
            filename = "students_template.xlsx"
            # Add note
            ws['A2'] = "Jane Doe"
            ws['B2'] = "jane.doe@example.com"
            ws['C2'] = "Female"
            ws['D2'] = "2010-05-15"
            ws['E2'] = "Class 10"
            ws['F2'] = "A"
            ws['G2'] = 101
            ws['H2'] = "John Doe"
            ws['I2'] = "Mary Doe"
            ws['J2'] = "1234567890"
            ws['K2'] = "2024-2025"

        elif template_type == 'subjects':
            headers = ['subject_name', 'subject_code']
            ws.title = "Subjects"
            filename = "subjects_template.xlsx"
        
        elif template_type == 'classes':
            headers = ['class_name', 'section', 'teacher_username']
            ws.title = "Classes"
            filename = "classes_template.xlsx"
        
        elif template_type == 'results':
            headers = [
                'student_username', 'subject_code', 'marks_obtained', 
                'total_marks', 'exam_type', 'academic_year'
            ]
            ws.title = "Results"
            filename = "results_template.xlsx"
        
        else:
            return None, None

        ws.append(headers)
        
        # Save to a byte stream
        file_data = io.BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        
        return file_data, filename

    @staticmethod
    def import_students(file_stream):
        """
        Imports students from an Excel file.
        This function calls the Student.create_student model method,
        which automatically creates the user account and student profile.
        """
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            wb = openpyxl.load_workbook(file_stream)
            ws = wb.active
            
            headers = [cell.value.lower().strip() for cell in ws[1]]
            required_headers = [
                'full_name', 'email', 'gender', 'date_of_birth', 
                'class_name', 'section', 'roll_number', 
                'fathers_name', 'mothers_name', 'academic_year'
            ]
            
            for header in required_headers:
                if header not in headers:
                    return False, f"Missing required column: {header}"
            
            added_count = 0
            failed_rows = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                data = dict(zip(headers, [cell.value for cell in row]))
                
                try:
                    # 1. Find class_id from class_name and section
                    class_name = data.get('class_name')
                    section = data.get('section')
                    
                    class_data = cursor.execute(
                        "SELECT id FROM classes WHERE class_name = ? AND section = ?",
                        (class_name, section)
                    ).fetchone()
                    
                    if not class_data:
                        failed_rows.append(f"Row {row_idx}: Class '{class_name} - {section}' not found.")
                        continue
                        
                    class_id = class_data['id']
                    
                    # 2. Get all required data
                    full_name = data.get('full_name')
                    email = data.get('email')
                    gender = data.get('gender')
                    # Ensure DOB is a string in YYYY-MM-DD format
                    date_of_birth_raw = data.get('date_of_birth')
                    if isinstance(date_of_birth_raw, datetime.datetime):
                         date_of_birth = date_of_birth_raw.strftime('%Y-%m-%d')
                    else:
                         date_of_birth = str(date_of_birth_raw).split(' ')[0]
                         
                    roll_number = int(data.get('roll_number'))
                    fathers_name = data.get('fathers_name')
                    mothers_name = data.get('mothers_name')
                    mobile_number = str(data.get('mobile_number', ''))
                    academic_year = data.get('academic_year')

                    # 3. Call the create_student model function
                    student_id, message = Student.create_student(
                        full_name, gender, date_of_birth, class_id, roll_number,
                        fathers_name, mobile_number, mothers_name, email, academic_year
                    )
                    
                    if student_id:
                        added_count += 1
                    else:
                        failed_rows.append(f"Row {row_idx} ({full_name}): {message}")
                        
                except Exception as e:
                    failed_rows.append(f"Row {row_idx}: Error processing - {str(e)}")

            conn.close()
            
            message = f"Import complete. Successfully added {added_count} students."
            if failed_rows:
                message += "\n\nFailed rows:\n" + "\n".join(failed_rows)
            
            return True, message

        except Exception as e:
            conn.close()
            return False, f"An error occurred: {str(e)}"

    @staticmethod
    def import_users(file_stream):
        """Imports Admin/Teacher users from an Excel file."""
        conn = get_db_connection()
        try:
            wb = openpyxl.load_workbook(file_stream)
            ws = wb.active
            headers = [cell.value.lower().strip() for cell in ws[1]]
            
            added_count = 0
            failed_rows = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                data = dict(zip(headers, [cell.value for cell in row]))
                
                username = data.get('username')
                password = data.get('password')
                role = data.get('role')
                name = data.get('name')
                email = data.get('email', '')
                
                if not (username and password and role and name):
                    failed_rows.append(f"Row {row_idx}: Missing required data.")
                    continue
                
                if role == 'student':
                    failed_rows.append(f"Row {row_idx} ({username}): Cannot bulk-add students. Use 'Upload Students' form.")
                    continue
                
                user_id = User.create_user(username, str(password), role, name, email)
                if user_id:
                    added_count += 1
                else:
                    failed_rows.append(f"Row {row_idx} ({username}): Username already exists.")
            
            conn.close()
            message = f"Import complete. Successfully added {added_count} users."
            if failed_rows:
                message += "\n\nFailed rows:\n" + "\n".join(failed_rows)
            return True, message
        except Exception as e:
            conn.close()
            return False, f"An error occurred: {str(e)}"

    @staticmethod
    def import_subjects(file_stream):
        """Imports subjects from an Excel file."""
        conn = get_db_connection()
        try:
            wb = openpyxl.load_workbook(file_stream)
            ws = wb.active
            headers = [cell.value.lower().strip() for cell in ws[1]]
            
            added_count = 0
            failed_rows = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                data = dict(zip(headers, [cell.value for cell in row]))
                
                name = data.get('subject_name')
                code = data.get('subject_code')
                
                if not (name and code):
                    failed_rows.append(f"Row {row_idx}: Missing name or code.")
                    continue
                
                if not Subject.create_subject(name, code):
                    failed_rows.append(f"Row {row_idx} ({name}): Subject name or code already exists.")
                else:
                    added_count += 1
            
            conn.close()
            message = f"Import complete. Successfully added {added_count} subjects."
            if failed_rows:
                message += "\n\nFailed rows:\n" + "\n".join(failed_rows)
            return True, message
        except Exception as e:
            conn.close()
            return False, f"An error occurred: {str(e)}"

    @staticmethod
    def import_classes(file_stream):
        """Imports classes from an Excel file."""
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            wb = openpyxl.load_workbook(file_stream)
            ws = wb.active
            headers = [cell.value.lower().strip() for cell in ws[1]]
            
            added_count = 0
            failed_rows = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                data = dict(zip(headers, [cell.value for cell in row]))
                
                class_name = data.get('class_name')
                section = data.get('section')
                teacher_username = data.get('teacher_username')
                
                if not (class_name and section):
                    failed_rows.append(f"Row {row_idx}: Missing class_name or section.")
                    continue
                
                # Find teacher_id from username
                teacher_id = None
                if teacher_username:
                    teacher = User.get_by_username(teacher_username)
                    if teacher and teacher.role == 'teacher':
                        teacher_id = teacher.id
                    else:
                        failed_rows.append(f"Row {row_idx}: Teacher '{teacher_username}' not found or is not a teacher.")
                        continue
                else:
                    failed_rows.append(f"Row {row_idx}: Missing teacher_username.")
                    continue

                class_id, message = Class.create_class(class_name, section, teacher_id)
                if class_id:
                    added_count += 1
                else:
                    failed_rows.append(f"Row {row_idx} ({class_name}): {message}")
            
            conn.close()
            message = f"Import complete. Successfully added {added_count} classes."
            if failed_rows:
                message += "\n\nFailed rows:\n" + "\n".join(failed_rows)
            return True, message
        except Exception as e:
            conn.close()
            return False, f"An error occurred: {str(e)}"

    @staticmethod
    def import_results(file_stream):
        """Imports results from an Excel file."""
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            wb = openpyxl.load_workbook(file_stream)
            ws = wb.active
            headers = [cell.value.lower().strip() for cell in ws[1]]
            
            added_count = 0
            failed_rows = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                data = dict(zip(headers, [cell.value for cell in row]))
                
                try:
                    student_username = data.get('student_username')
                    subject_code = data.get('subject_code')
                    marks_obtained = float(data.get('marks_obtained'))
                    total_marks = float(data.get('total_marks'))
                    exam_type = data.get('exam_type')
                    academic_year = data.get('academic_year')
                    
                    # Find student_id
                    student = User.get_by_username(student_username)
                    if not (student and student.role == 'student'):
                        failed_rows.append(f"Row {row_idx}: Student user '{student_username}' not found.")
                        continue
                    
                    # Find subject_id
                    subject = cursor.execute("SELECT id FROM subjects WHERE subject_code = ?", (subject_code,)).fetchone()
                    if not subject:
                        failed_rows.append(f"Row {row_idx}: Subject code '{subject_code}' not found.")
                        continue
                    
                    result_id, message = Result.enter_marks(
                        student.id, subject['id'], marks_obtained, 
                        total_marks, exam_type, academic_year
                    )
                    
                    if result_id:
                        added_count += 1
                    else:
                        failed_rows.append(f"Row {row_idx} ({student_username}): {message}")

                except Exception as e:
                    failed_rows.append(f"Row {row_idx}: Error processing - {str(e)}")
            
            conn.close()
            message = f"Import complete. Successfully added {added_count} results."
            if failed_rows:
                message += "\n\nFailed rows:\n" + "\n".join(failed_rows)
            return True, message
        except Exception as e:
            conn.close()
            return False, f"An error occurred: {str(e)}"